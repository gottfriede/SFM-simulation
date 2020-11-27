import openpyxl
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D

plt.rcParams['font.sans-serif'] = ['KaiTi'] # 指定默认字体
plt.rcParams['axes.unicode_minus'] = False # 解决保存图像是负号'-'显示为方块的问题

if __name__ == '__main__' :
    x = []
    y = []
    z = []
    for i in [10,20,30,40] :
        for j in range(4) :
            x.append(i)
    for i in range(4) :
        for j in [1,2,3,4] :
            y.append(j)

    workbook = load_workbook('仿真结果.xlsx')
    worksheet = workbook['Sheet1']
    i = 10
    while i <= 37 :
        z.append(float(worksheet.cell(i, 5).value))
        z.append(float(worksheet.cell(i+1, 5).value))
        z.append(float(worksheet.cell(i+2, 5).value))
        z.append(float(worksheet.cell(i+3, 5).value))
        i = i + 8
    
    # print(x)
    # print(y)
    # print(z)

    fig = plt.figure()
    ax = fig.gca(projection='3d')
 
    ax.plot_trisurf(x, y, z, linewidth=0.2, antialiased=True)
    ax.set_xlabel('房间人数')
    ax.set_ylabel('出口宽度')
    ax.set_zlabel('撤离时间')
 
    plt.show()
